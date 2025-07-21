def run():
    #This script loads evaluatie template data from excel file.
    # Prerequisite: Schoolgroepn naam moet eerst opgezet zijn in de database en moet matchen met waarde in sheet EvaluatieTemplates
    import openpyxl
    import os
    import django
    import datetime as dt

    os.environ.setdefault("DJANGO_SETTINGS_MODULE","stagetoolproj.settings")
    from django.db import models
    from stageapp.models import EvaluatieCriteria, EvaluatieScoreRange, EvaluatieTemplateDetails, EvaluatieTemplateHoofddeel, EvaluatieTemplateSubdeel, EvaluatieTemplates, ScoreType, ScholenGroep
    from django.contrib.auth.models import User #To restrict access to the user who created the entry
    django.setup()


    strpf = 'D:/Privé/UploadInitialData.xlsx'
    datecreated = dt.datetime.now().date()

    wrkbk = openpyxl.load_workbook(filename=strpf)
    shtnames = wrkbk.sheetnames
    #print(shtnames)
    for sheet in wrkbk.worksheets:
        #print(sheet.title)
        for row in sheet.iter_rows(values_only=True):  #Values only is required to really capture the content of the cells
            #print(row)
            #print(f"     {row[0]}")
            if sheet.title == 'EvaluatieTemplateHoofddeel':
                hoofddeel = row[0]
                hoofddeel_omschrijving = row[1]
                hd = EvaluatieTemplateHoofddeel(
                        hoofddeel = hoofddeel,
                        hoofddeel_omschrijving = hoofddeel_omschrijving,
                        created_by=User.objects.get(username='Admin'),
                        created_at = datecreated,
                        last_updated_at = datecreated,
                        last_updated_by = User.objects.get(username='Admin'),
                )
                hd.save()
            if sheet.title == 'EvaluatieTemplates':
                sg = row[0]
                template_naam = row[1]
                template_naam = row[1]
                et = EvaluatieTemplates(
                    scholengroep = ScholenGroep.objects.get(scholen_groep_naam=sg),
                    template_naam=template_naam,
                    template_omschrijving=template_naam,
                    created_by=User.objects.get(username='Admin'),
                    created_at=datecreated,
                    last_updated_at=datecreated,
                    last_updated_by=User.objects.get(username='Admin'),
                )
                et.save()
            if sheet.title == 'EvaluatieTemplateSubdeel':
                subdeel = row[0]
                subdeel_omschrijving = row[1]
                sd = EvaluatieTemplateSubdeel(
                    subdeel=subdeel,
                    subdeel_omschrijving=subdeel_omschrijving,
                    created_by=User.objects.get(username='Admin'),
                    created_at=datecreated,
                    last_updated_at=datecreated,
                    last_updated_by=User.objects.get(username='Admin'),
                )
                sd.save()
            if sheet.title == 'EvaluatieCriteria':
                criteria = row[0]
                criteria_omschrijving = row[0]
                ec = EvaluatieCriteria(
                    criteria=criteria,
                    criteria_omschrijving=criteria,
                    created_by=User.objects.get(username='Admin'),
                    created_at=datecreated,
                    last_updated_at=datecreated,
                    last_updated_by=User.objects.get(username='Admin'),
                )
                ec.save()
            if sheet.title == 'ScoreType':
                #print(row)
                scoretype = row[0]
                scoretype_omschrijving = row[1]
                st = ScoreType(
                    score_type=scoretype,
                    score_type_omschrijving = scoretype_omschrijving,
                    created_by=User.objects.get(username='Admin'),
                    created_at=datecreated,
                    last_updated_at=datecreated,
                    last_updated_by=User.objects.get(username='Admin'),
                )
                st.save()
            if sheet.title == 'EvaluatieScoreRange':
                scoretype = row[0]
                score_range = row[1]
                score_van = row[2]
                score_tot = row[3]
                toelichting_score_range = row[4]
                esr = EvaluatieScoreRange(
                    scoretype=ScoreType.objects.get(score_type=scoretype),
                    score_range = score_range,
                    score_van = score_van,
                    score_tot = score_tot,
                    toelichting_score_range = toelichting_score_range,
                    created_by=User.objects.get(username='Admin'),
                    created_at=datecreated,
                    last_updated_at=datecreated,
                    last_updated_by=User.objects.get(username='Admin'),
                )
                esr.save()
            if sheet.title == 'EvaluatieTemplateDetails':
                evaluatietemplates = row[0]
                hoofddeel = row[1]
                subdeel = row[2]
                criteria = row[3]
                score_range = row[4]
                if row[2]:  #This to check if subdeel is entered.  If so, then it continue, else blank and a blank subdeel is entered
                    #print("different from none")
                    etd = EvaluatieTemplateDetails(
                        evaluatietemplates = EvaluatieTemplates.objects.get(template_naam= evaluatietemplates),
                        hoofddeel = EvaluatieTemplateHoofddeel.objects.get(hoofddeel=hoofddeel),
                        subdeel = EvaluatieTemplateSubdeel.objects.get(subdeel = subdeel),
                        criteria = EvaluatieCriteria.objects.get(criteria = criteria),
                        score_range = EvaluatieScoreRange.objects.get(score_range = score_range),
                        created_by=User.objects.get(username='Admin'),
                        created_at=datecreated,
                        last_updated_at=datecreated,
                        last_updated_by=User.objects.get(username='Admin'),
                    )
                    etd.save()
                else:
                    #print("equal to none")
                    etd = EvaluatieTemplateDetails(
                        evaluatietemplates = EvaluatieTemplates.objects.get(template_naam=evaluatietemplates),
                        hoofddeel = EvaluatieTemplateHoofddeel.objects.get(hoofddeel=hoofddeel),
                        subdeel = None,
                        criteria = EvaluatieCriteria.objects.get(criteria=criteria),
                        score_range = EvaluatieScoreRange.objects.get(score_range=score_range),
                        created_by = User.objects.get(username='Admin'),
                        created_at = datecreated,
                        last_updated_at = datecreated,
                        last_updated_by = User.objects.get(username='Admin'),
                        )
                    etd.save()